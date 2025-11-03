#****************************************  Extractor XML   *******************************************************
# EricK Genaro Piña Arrieta Barbosa
# Proyecto: Extractor de datos XML en Python
# Fecha: 31/10/2025
# Descripción: Este proyecto tiene como objetivo extraer y procesar datos de un XML CFDI 4 utilizando Python.
#*****************************************************************************************************************

import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import os

#-------------------------------------------  Función para definir ruta  ----------------------------------------#
def definir_ruta_archivos():
    print('\nProbablemente tus archivos se encuentran en la carpeta de Descargas')
    print('Ejemplo en Windows: C:/Users/TuUsuario/Downloads/CFDI00000000.xml')
    print('Ejemplo en Linux Ubuntu: /home/TuUsuario/Descargas/CFDI00000000.xml\n')
    ruta = input('Ingrese la ruta completa del archivo XML: ').strip()

    if not os.path.exists(ruta):
        print('\nRuta inválida o archivo no encontrado. Inténtalo de nuevo.\n')
        return definir_ruta_archivos()
    
    print('\nVerifica si la ruta y el nombre del archivo son correctos.')
    input('Presiona ENTER para continuar ...\n')
    return ruta


#-------------------------------------------  Función para extraer UUID  ----------------------------------------#
def extraer_uuid(root, namespaces):
    """Busca el UUID dentro del nodo <tfd:TimbreFiscalDigital>."""
    complemento = root.find('cfdi:Complemento', namespaces)
    if complemento is not None:
        timbre = complemento.find('tfd:TimbreFiscalDigital', namespaces)
        if timbre is not None:
            return timbre.attrib.get('UUID')
    return None


#-------------------------------------------  Función para parsear XML  -----------------------------------------#
def parseXML():
    xml_file = definir_ruta_archivos()
    namespaces = {
        'cfdi': 'http://www.sat.gob.mx/cfd/4',
        'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'
    }

    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
    except ET.ParseError:
        print("\nError: El archivo XML está dañado o no tiene el formato esperado.")
        return None

    # Datos generales
    serie = root.attrib.get('Serie')
    folio = root.attrib.get('Folio')
    fecha = root.attrib.get('Fecha')
    subtotal = root.attrib.get('SubTotal')
    descuento = root.attrib.get('Descuento')
    total = root.attrib.get('Total')

    # Nodos principales
    emisor = root.find('cfdi:Emisor', namespaces)
    receptor = root.find('cfdi:Receptor', namespaces)
    impuestos = root.find('cfdi:Impuestos', namespaces)

    totalImpuestosTrasladados = None
    if impuestos is not None:
        totalImpuestosTrasladados = impuestos.attrib.get('TotalImpuestosTrasladados')

    # UUID del timbre fiscal
    uuid = extraer_uuid(root, namespaces)

    print('\nDatos extraídos del CFDI:')
    print("UUID:", uuid)
    print("Serie:", serie)
    print("Folio:", folio)
    print("Fecha:", fecha)
    print("SubTotal:", subtotal)
    print("Descuento:", descuento)
    print("Total_Impuestos_trasladados:", totalImpuestosTrasladados)
    print("Total:", total)

    RFC_Emisor = Nombre_Emisor = RFC_Receptor = Nombre_Receptor = None

    if emisor is not None:
        RFC_Emisor = emisor.attrib.get('Rfc')
        Nombre_Emisor = emisor.attrib.get('Nombre')
        print("RFC_Emisor:", RFC_Emisor)
        print("Nombre_Emisor:", Nombre_Emisor)

    if receptor is not None:
        RFC_Receptor = receptor.attrib.get('Rfc')
        Nombre_Receptor = receptor.attrib.get('Nombre')
        print("RFC_Receptor:", RFC_Receptor)
        print("Nombre_Receptor:", Nombre_Receptor)

    # Conversión segura a float
    def safe_float(val):
        try:
            return float(str(val).replace(',', '.')) if val else 0.0
        except ValueError:
            return 0.0

    miFila = [
        uuid, serie, folio, fecha,
        safe_float(subtotal),
        safe_float(descuento),
        safe_float(totalImpuestosTrasladados),
        safe_float(total),
        RFC_Emisor, Nombre_Emisor,
        RFC_Receptor, Nombre_Receptor
    ]

    return miFila


#-------------------------------------------  Crear Excel con encabezados  --------------------------------------#
def crearExcelconEncabezados(nombreXLS):
    encabezados = [
        'UUID', 'Serie', 'Folio', 'Fecha', 'SubTotal', 'Descuento',
        'Total_Impuestos_trasladados', 'Total',
        'RFC_Emisor', 'Nombre_Emisor', 'RFC_Receptor', 'Nombre_Receptor'
    ]

    df = pd.DataFrame(columns=encabezados)
    print('\nVista previa del dataframe vacío:')
    print(df)

    ruta = input('\nDefine la ruta donde se guardará el archivo (termina con / o \\): ').strip()
    if not os.path.isdir(ruta):
        print('La ruta no existe. Intenta de nuevo.')
        return crearExcelconEncabezados(nombreXLS)
    
    rutaXLS = os.path.join(ruta, nombreXLS)
    df.to_excel(rutaXLS, index=False)
    return rutaXLS


#-------------------------------------------  Convertir lista a diccionario  ------------------------------------#
def convertir_lista_a_diccionario(registro):
    claves = [
        'UUID', 'Serie', 'Folio', 'Fecha', 'SubTotal', 'Descuento',
        'Total_Impuestos_trasladados', 'Total',
        'RFC_Emisor', 'Nombre_Emisor', 'RFC_Receptor', 'Nombre_Receptor'
    ]
    return dict(zip(claves, registro))


#-------------------------------------------  Insertar fila en Excel  -------------------------------------------#
def insertarFilaCFDI(ruta_archivo, diccionario):
    df_nueva = pd.DataFrame([diccionario])
    libro = load_workbook(ruta_archivo)
    hoja = libro.sheetnames[0]
    hoja_existente = pd.read_excel(ruta_archivo, sheet_name=hoja)

    with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        startrow = len(hoja_existente) + 1
        df_nueva.to_excel(writer, sheet_name=hoja, index=False, header=False, startrow=startrow)

    print("\nFila agregada exitosamente.\n")


#-------------------------------------------  Programa principal  -----------------------------------------------#
if __name__ == "__main__":
    print('\n**********************************************************************************')
    print('************************** Extractor de datos XML CFDI 4.0 *************************')
    print('**********************************************************************************\n')

    print('A continuación deberás indicar un nombre y una ruta para tu archivo Excel:')
    nombreXLS = input('Nombre del archivo (sin extensión): ').strip() + '.xlsx'
    pathXLS = crearExcelconEncabezados(nombreXLS)
    print(f'\nEl archivo fue creado correctamente en: {pathXLS}\n')
    input('Presiona ENTER para continuar y cargar tus CFDI...')

    contador = 0
    while True:
        registro = parseXML()
        if registro is None:
            print('No se pudo procesar el XML. Intenta con otro archivo.')
        else:
            print('\nLos datos extraídos son:\n', registro)
            elDiccionario = convertir_lista_a_diccionario(registro)
            insertarFilaCFDI(pathXLS, elDiccionario)
            contador += 1

        continuar = input('¿Deseas agregar otro CFDI? (s/n): ').strip().lower()
        if continuar != 's':
            break

    print(f'\nSe procesaron correctamente {contador} archivo(s).')
    print('\n*************************************************************************************************************')
    print('************************************** Fin del programa *****************************************************')
    print('*************************************************************************************************************')
    print('Gracias por usar la aplicación. Cualquier comentario a: pinaarrieta@yahoo.com.mx\n')
    input('Presiona ENTER para finalizar...')